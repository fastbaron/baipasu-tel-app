import shutil
from flask import Flask, render_template, request, redirect, url_for
import sqlite3
import pandas as pd
from collections import defaultdict
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from datetime import datetime


app = Flask(__name__)


class Client:
    def __init__(self, id, name, course, items):
        self.id = id
        self.name = name
        self.course = course
        self.items = items  # ここを修正


def load_clients_from_file(file_path, file_type="csv"):
    if file_type == "csv":
        df = pd.read_csv(file_path)
    elif file_type == "excel":
        df = pd.read_excel(file_path)
    else:
        raise ValueError("Invalid file type. Use 'csv' or 'excel'.")

    clients = []
    unique_clients = df.drop_duplicates(subset=["id", "name", "course"])

    for index, row in unique_clients.iterrows():
        client_id = row["id"]
        client_name = row["name"]
        course = row["course"]
        items = [
            {"name": item_row["item1"]}
            for _, item_row in df[df["id"] == client_id].iterrows()
        ]
        clients.append(Client(client_id, client_name, course, items))

    return clients


clients = load_clients_from_file(
    "clients_2.xlsx", file_type="excel"
)  # Use "clients.xlsx" and file_type="excel" for Excel file


def get_client(id):
    for client in clients:
        if client.id == id:
            return client
    return None


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/id_submit", methods=["POST"])
def id_submit():
    client_id = request.form.get("client_id")
    if client_id and get_client(int(client_id)):
        return redirect(url_for("client_page", id=client_id))
    else:
        return "Client not found", 404


@app.route("/client/<int:id>", methods=["GET"])
def client_page(id):
    client = get_client(id)
    if client:
        return render_template("client.html", client=client)
    else:
        return "Client not found", 404


@app.route("/submit", methods=["POST"])
def submit_order():
    client_id = int(request.form["client_id"])  # Convert to int
    client_name = request.form["client_name"]
    course = request.form["course"]
    items = {
        key.split("_")[1]: int(value)
        for key, value in request.form.items()
        if key not in ["client_id", "client_name", "course"]
    }

    order = {
        "client_id": client_id,  # Use int value here
        "client_name": client_name,
        "course": course,
        "items": items,
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }

    print(order)  # Add this line to print the order

    save_order_to_excel(order)  # Pass order here

    return render_template("order_success.html")


def save_order_to_excel(order):
    current_time = datetime.now()
    filename = f"orders_{current_time.strftime('%Y%m%d')}.xlsx"
    file_exists = os.path.isfile(filename)

    if not file_exists:
        shutil.copy("clients_2.xlsx", filename)

    wb = load_workbook(filename)
    ws = wb.active

    for item, quantity in order["items"].items():
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=5):
            if (
                row[0].value == order["client_id"]  # No need to convert here
                and row[1].value == int(order["course"])
                and row[3].value == item
            ):
                row[4].value = quantity
                print(
                    f"Updated cell: {row[4].coordinate}"
                )  # Add this line to print the updated cell coordinate

                break

    wb.save(filename)


if __name__ == "__main__":
    # app.run(debug=True)
    app.run(host='0.0.0.0', debug=True)
