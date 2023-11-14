import os
import time
import pandas as pd
import gspread
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from openpyxl import load_workbook
import gspread_dataframe

# サービスアカウントキーで認証
credentials = service_account.Credentials.from_service_account_file(
    "jyutyuu-7338f82eace1.json"
)

# スコープを設定
scoped_credentials = credentials.with_scopes(
    [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
)

# Google Sheets APIとGoogle Drive APIクライアントを作成
sheets_api = build("sheets", "v4", credentials=scoped_credentials)
drive_api = build("drive", "v3", credentials=scoped_credentials)

# Google Drive上のスプレッドシートIDとローカルのExcelファイルのパスを指定
spreadsheet_id = "1Wk7geXlFOlygvk0fMNPop9WeQLr3RkXdHzUyQBHJrBc"
excel_file_path = (
    "C:\\Users\\junim\\OneDrive - ㈱バイパス給食センター\\myFolder_5\\orders_20230420.xlsx"
)


def upload_data():
    # ExcelファイルをPandas DataFrameに読み込む
    wb = load_workbook(excel_file_path)
    sheet_name = wb.sheetnames[0]
    df = pd.read_excel(excel_file_path, sheet_name=sheet_name)

    # # Googleスプレッドシートをクリア
    # sheets_api.spreadsheets().values().clear(
    #     spreadsheetId=spreadsheet_id, range=sheet_name, body={}
    # ).execute()

    # DataFrameをスプレッドシートに書き込む
    gc = gspread.authorize(scoped_credentials)
    sh = gc.open_by_key(spreadsheet_id)
    worksheet = sh.get_worksheet(0)
    gspread_dataframe.set_with_dataframe(worksheet, df)


while True:
    try:
        upload_data()
        print("データをアップロードしました")
    except HttpError as error:
        print(f"エラーが発生しました: {error}")
    time.sleep(3)  # 3秒待機
